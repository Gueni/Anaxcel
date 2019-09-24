import functools
import os
import random
import sys
import time
import matplotlib.pyplot as plt
import openpyxl
import pandas as pd
import xlrd
import xlwt
from PyQt5 import QtGui, QtWidgets, QtCore
from PyQt5.QtChart import QChart, QChartView, QPieSeries, QPieSlice
from PyQt5.QtCore import Qt, QTimer, QCoreApplication, QSettings
from PyQt5.QtGui import QPixmap, QPainter
from PyQt5.QtWidgets import QFileDialog, QTableWidgetItem, QGridLayout, QSplashScreen
from PyQt5.QtWidgets import QVBoxLayout
from matplotlib.backends.backend_qt5agg import FigureCanvasQTAgg as FigureCanvas
from matplotlib.backends.backend_qt5agg import NavigationToolbar2QT as NavigationToolbar
from matplotlib.ticker import PercentFormatter
import anaxcelgui
import aboutmain





class anaxcelhandler(QtWidgets.QMainWindow, anaxcelgui.Ui_MainWindow):

    def __init__(self, parent=None):
        super(anaxcelhandler, self).__init__(parent)
        if getattr(sys, 'frozen', False):
            self.frozen = 'ever so'
            self.bundle_dir = sys._MEIPASS
        else:
            self.bundle_dir = os.path.dirname(os.path.abspath(__file__))
        self.setupUi(self)
        self.setWindowIcon(QtGui.QIcon(self.bundle_dir + '/icons/icon.png'))
        self.setStyleSheet(open("Dark/darkstyle.qss", "r").read())
        self.listWidget.setSelectionMode(QtWidgets.QAbstractItemView.ExtendedSelection)
        self.pushButtonbrowse.clicked.connect(self.openFileNamesDialog)
        self.pushButtonclear.clicked.connect(self.clearwidget)
        self.pushButtonselall.clicked.connect(self.selectall)
        self.pushButtonload.clicked.connect(self.LoadProcess)
        self.statusbar.showMessage('V 0.1')
        self.actionExit.setShortcut('Ctrl+Q')
        self.actionExit.triggered.connect(self.close_application)
        self.actionExit.setStatusTip('Exit ')
        self.pixmaplab = QPixmap('icons/excel48.png')
        self.labelicon.setPixmap(self.pixmaplab)
        self.labelicon.show()
        self.label.setText("Loaded Excel File")
        self.comboBoxfiletype.addItems(['xls', ' xlsx'])
        self.pushButton_5.clicked.connect(self.analyseProcess)
        self.figure = plt.figure()
        self.canvas = FigureCanvas(self.figure)
        self.toolbar = NavigationToolbar(self.canvas, self)
        layoutx = QVBoxLayout()
        layoutx.addWidget(self.toolbar)
        layoutx.addWidget(self.canvas)
        self.tab_6.setLayout(layoutx)
        self.layoutdout = QGridLayout()
        self.tab_4.setLayout(self.layoutdout)
        self.chartView3 = QChartView()
        self.layoutdout.addWidget(self.chartView3)
        self.actionAdd_File.setShortcut('Ctrl+A')
        self.actionAdd_File.triggered.connect(self.openFileNamesDialog)
        self.actionAdd_File.setStatusTip('Add File ')

        self.actionClear.setShortcut('Ctrl+C')
        self.actionClear.triggered.connect(self.clearwidget)
        self.actionClear.setStatusTip('Clear ')

        self.actionSelect_All.setShortcut('Ctrl+S')
        self.actionSelect_All.triggered.connect(self.selectall)
        self.actionSelect_All.setStatusTip('Sellect All ')

        self.actionLoad.setShortcut('Ctrl+L')
        self.actionLoad.triggered.connect(self.LoadProcess)
        self.actionLoad.setStatusTip('Load ')

        self.actionAbout.setShortcut('Ctrl+I')
        self.actionAbout.triggered.connect(self.showAbout)
        self.actionAbout.setStatusTip('About ')




    def showAbout(self):
        self.window = aboutmain.abouthandel()
        self.window.show()

    def close_application(self):
        choice = QtWidgets.QMessageBox.question(self, ' Confirm Exit ', "Are You Sure You want To Close Anaxcel ?",
                                                QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No)
        if choice == QtWidgets.QMessageBox.Yes:
            self.saveSettingstofilepackage()
            sys.exit()
        else:
            pass

    def openFileNamesDialog(self):
        options = QFileDialog.Options()
        options |= QFileDialog.DontUseNativeDialog
        filterxls = "XLS (*.xls *.XLS)"
        filterxlsx = "XLSX (*.xlsx *.XLSX)"
        if self.comboBoxfiletype.currentIndex() == 0:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLS Files", filter=filterxls, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)
        elif self.comboBoxfiletype.currentIndex() == 1:
            files, _ = QFileDialog.getOpenFileNames(self, "Select XLSX Files", filter=filterxlsx, options=options)
            if files:
                for file in files:
                    self.listWidget.addItem(file)

    def clearwidget(self):
        self.listWidget.clear()
        self.tableWidget.clear()
        self.label.setText("Loaded Excel File")

    def selectall(self):
        self.listWidget.selectAll()

    # def CsvProcess(self):
    # # merge csv files
    # fout = open("combined.csv", "a")
    # items = self.listWidget.selectedItems()
    # xlsfiles = []
    # for i in list(items):
    #     xlsfiles.append(str(i.text()))
    # for line in open(xlsfiles[0]):
    #     fout.write(line)
    # for index in xlsfiles[1:]:
    #     print(index)
    #     # now the rest:
    #     f = open(index)
    #     f.__next__()  # skip the header
    #     for line in f:
    #         fout.write(line)
    #     f.close()  # not really needed
    # fout.close()

    # f = open("folder/02-08-2017.CSV")
    # for line in f:
    #     for x in line:
    #         if x == ';'
    #     line.split(';')
    #     print(line)
    #
    # f.close()  # not really needed

    def xlsProcess(self):
        self.tableWidget.clear()
        items = self.listWidget.selectedItems()
        xlsfiles = []
        for i in list(items):
            xlsfiles.append(i.text())
        wkbk = xlwt.Workbook()
        outsheet = wkbk.add_sheet('Sheet1')
        outrow_idx = 0
        for f in xlsfiles:
            insheet = xlrd.open_workbook(f).sheets()[0]
            for row_idx in range(insheet.nrows):
                for col_idx in range(insheet.ncols):
                    outsheet.write(outrow_idx, col_idx, insheet.cell_value(row_idx, col_idx))
                outrow_idx += 1
        wkbk.save(r'combined.xls')
        # use on_demand=True to avoid loading worksheet data into memory
        book = xlrd.open_workbook("combined.xls", on_demand=True)
        sheet = book.sheet_by_index(0)
        num_rows = sheet.nrows
        num_col = sheet.ncols
        self.tableWidget.setRowCount(num_rows)
        self.tableWidget.setColumnCount(num_col)
        for col in range(num_col):
            for row in range(num_rows):
                cell = sheet.cell(row, col)
                if (not cell.value == "") and (not cell.value == " "):
                    self.tableWidget.setItem(row, col, QTableWidgetItem(str(cell.value)))
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()

    def xlsxprocess(self):
        self.tableWidget.clear()
        items = self.listWidget.selectedItems()
        xlsfiles = []
        xlsoutput = []
        c = 0
        for i in list(items):
            xlsfiles.append(str(i.text()))
        for xlsx_item in xlsfiles:
            wb = openpyxl.load_workbook(xlsx_item)
            pathname = "file" + str(c) + ".xls"
            wb.save(pathname)
            c += 1
            xlsoutput.append(str(pathname))
        wkbk = xlwt.Workbook()
        outsheet = wkbk.add_sheet('Sheet1')
        outrow_idx = 0
        for f in xlsoutput:
            print(f)
            insheet = xlrd.open_workbook(f).sheets()[0]
            for row_idx in range(insheet.nrows):
                for col_idx in range(insheet.ncols):
                    outsheet.write(outrow_idx, col_idx, insheet.cell_value(row_idx, col_idx))
                outrow_idx += 1
        wkbk.save(r'combined.xls')
        df = pd.read_excel('combined.xls', header=0)  # read file and set header row
        self.tableWidget.setColumnCount(len(df.columns))
        self.tableWidget.setRowCount(len(df.index))
        for i in range(len(df.index)):
            for j in range(len(df.columns)):
                self.tableWidget.setItem(i, j, QTableWidgetItem(str(df.iat[i, j])))
        self.tableWidget.resizeColumnsToContents()
        self.tableWidget.resizeRowsToContents()

    def LoadProcess(self):
        if self.comboBoxfiletype.currentIndex() == 0:  # xls
            self.xlsProcess()
            for colindex in range(self.tableWidget.columnCount()):
                self.comboBoxabsc.addItem(str(colindex))
                self.comboBoxordo.addItem(str(colindex))
        elif self.comboBoxfiletype.currentIndex() == 1:  # xlsx
            self.xlsxprocess()
            for colindex in range(self.tableWidget.columnCount()):
                self.comboBoxabsc.addItem(str(colindex))
                self.comboBoxordo.addItem(str(colindex))

    def analyseProcess(self):
        absc = []
        ordo = []
        absc.clear()
        ordo.clear()
        for tabrow in range(self.tableWidget.rowCount()):
            abscitem = self.tableWidget.item(tabrow, self.comboBoxabsc.currentIndex())
            ordoitem = self.tableWidget.item(tabrow, self.comboBoxordo.currentIndex())
            absc.append(int(abscitem.text()))
            ordo.append(str(ordoitem.text()))
        df = pd.DataFrame({'data': absc})
        df.index = ordo
        df = df.sort_values(by='data', ascending=False)
        df["cumpercentage"] = df["data"].cumsum() / df["data"].sum() * 100
        self.figure.clear()
        plt.ion()
        ax = self.figure.add_subplot()
        ax.bar(df.index, df["data"], color="C0")
        ax2 = ax.twinx()
        ax2.plot(df.index, df["cumpercentage"], color="C1", marker="D", ms=7)
        ax2.yaxis.set_major_formatter(PercentFormatter())
        ax.tick_params(axis="y", colors="C0")
        ax2.tick_params(axis="y", colors="C1")
        self.canvas.draw()
        # donutchart***********************************
        self.m_donuts = []
        self.chartView3.setRenderHint(QPainter.Antialiasing)
        self.chart3 = self.chartView3.chart()
        self.chart3.legend().setVisible(True)
        self.chart3.setTitle("Nested donuts Chart")
        self.chart3.setAnimationOptions(QChart.AllAnimations)
        minSize3 = 0.1
        maxSize3 = 0.9
        donutCount3 = 5
        for i in range(donutCount3):
            donut = QPieSeries()
            sliceCount = random.randrange(3, 6)
            # print(sliceCount)
            for j in range(sliceCount):
                value3 = random.randrange(0, 50)
                slice_ = QPieSlice(str(value3), value3)
                slice_.setLabelVisible(True)
                slice_.setLabelColor(Qt.white)
                slice_.setLabelPosition(QPieSlice.LabelInsideTangential)
                slice_.hovered[bool].connect(functools.partial(self.explodeSlice, slice_=slice_))
                donut.append(slice_)
                donut.setHoleSize(minSize3 + i * (maxSize3 - minSize3) / donutCount3)
                donut.setPieSize(minSize3 + (i + 1) * (maxSize3 - minSize3) / donutCount3)
            self.m_donuts.append(donut)
            self.chartView3.chart().addSeries(donut)
        self.updateTimer = QTimer(self)
        self.updateTimer.timeout.connect(self.updateRotation)
        self.updateTimer.start(1250)
        self.tabWidget.setCurrentIndex(2)

    def updateRotation(self):
        for donut in self.m_donuts:
            phaseShift = random.randrange(-50, 100)
            donut.setPieStartAngle(donut.pieStartAngle() + phaseShift)
            donut.setPieEndAngle(donut.pieEndAngle() + phaseShift)

    def explodeSlice(self, exploded, slice_):
        if exploded:
            self.updateTimer.stop()
            sliceStartAngle = slice_.startAngle()
            sliceEndAngle = slice_.startAngle() + slice_.angleSpan()
            donut = slice_.series()
            seriesIndex = self.m_donuts.index(donut)
            for i in range(seriesIndex + 1, len(self.m_donuts)):
                self.m_donuts[i].setPieStartAngle(sliceEndAngle)
                self.m_donuts[i].setPieEndAngle(360 + sliceStartAngle)
        else:
            for donut in self.m_donuts:
                donut.setPieStartAngle(0)
                donut.setPieEndAngle(360)
            self.updateTimer.start()
        slice_.setExploded(exploded)


def main():
    app = QtWidgets.QApplication(sys.argv)
    # splash_pix = QPixmap('icons/splash.png')
    # splash = QSplashScreen(splash_pix, Qt.WindowStaysOnTopHint)
    # splash.setMask(splash_pix.mask())
    # splash.show()
    # time.sleep(2)
    # splash.close()
    window = anaxcelhandler()
    window.show()
    sys.exit(app.exec_())


if __name__ == '__main__':
    main()
